from time import sleep
from random import randint
import requests
import openpyxl
from aiogram import types
from bs4 import BeautifulSoup
from parsers.classes.proxies import ProxyManager

PROX = ProxyManager()

def request_decorator(func):
    def wrapper(f, **kwargs):
        a = True
        i = 1
        while a:
            response = func(f, **kwargs)
            print(response.status_code)
            if response.status_code == 200:
                sleep(randint(1, 2))
                a = False
            else:
                PROX.next()
            if i > 30:
                with open(f'{response.status_code}.html', 'w', encoding='utf8') as file:
                    file.write(response.text)
                raise Exception("Proxy don't work!")
            
            i += 1
        return response
    return wrapper


@request_decorator
def req(func, **kwargs):
    resp = func(**kwargs)
    return resp


def create_session():
    session = requests.Session()
    cookies = {
        '__cf_bm': 'YoQbzlehtc7L7vOOPgDdkFeNBbFzq2PFIIT9KyMc.SQ-1677365103-0-AYvVPBLOAqD3ewOdLqq+zqsMcSmHTp3u24aLjg6A8/kIPvG72DyRn3BRnktAl4hlZUD+K70dE7helMD2hEMRIRQ=',
    }

    headers = {
        'accept': '*/*',
        # 'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'en-US',
        'apollographql-client-name': 'Iron',
        'apollographql-client-version': '2023.02.12.02',
        'app-platform': 'Iron',
        'app-version': '2023.02.12.02',
        # 'content-length': '1402',
        'content-type': 'application/json',
        'origin': 'https://stockx.com',
        'referer': 'https://stockx.com/',
        'sec-ch-ua': '"Chromium";v="110", "Not A(Brand";v="24", "Google Chrome";v="110"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'selected-country': 'TH',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
        'x-operation-name': 'GetSearchResults',
        'x-stockx-device-id': '0e1a3324-7f22-47f9-a552-f9eec36bba6c',
        # 'Cookie': '__cf_bm=YoQbzlehtc7L7vOOPgDdkFeNBbFzq2PFIIT9KyMc.SQ-1677365103-0-AYvVPBLOAqD3ewOdLqq+zqsMcSmHTp3u24aLjg6A8/kIPvG72DyRn3BRnktAl4hlZUD+K70dE7helMD2hEMRIRQ=',
    }

    session.headers.update(headers)
    session.cookies.update(cookies)
    return session


def get_sb(filename: str):
    book = openpyxl.load_workbook(filename=filename)
    sheet = book["data"]
    return sheet, book


async def find_goods(filename: str, msg: types.Message):
    session = create_session()
    sheet, book = get_sb(filename)
    await msg.edit_text(f'Выполнено 0%. (0/{sheet.max_row-1})')
    for row in range (2, sheet.max_row+1):
        scrapSKU = (sheet[row][0].value)
        name = (sheet[row][1].value)
        print("Взял артикул " + str(scrapSKU))
        if name is None:
            json_data = {
                'query': 'query GetSearchResults($filtersVersion: Int, $query: String!, $page: BrowsePageInput, $sort: BrowseSortInput, $staticRanking: BrowseExperimentStaticRankingInput) {\n  browse(\n    query: $query\n    page: $page\n    sort: $sort\n    filtersVersion: $filtersVersion\n    experiments: {staticRanking: $staticRanking}\n  ) {\n    categories {\n      id\n      name\n      count\n    }\n    results {\n      edges {\n        objectId\n        node {\n          ... on Product {\n            id\n            urlKey\n            primaryTitle\n            secondaryTitle\n            media {\n              thumbUrl\n            }\n            brand\n            productCategory\n          }\n          ... on Variant {\n            id\n            product {\n              id\n              urlKey\n              primaryTitle\n              secondaryTitle\n              media {\n                thumbUrl\n              }\n              brand\n              productCategory\n            }\n          }\n        }\n      }\n      pageInfo {\n        limit\n        page\n        pageCount\n        queryId\n        queryIndex\n        total\n      }\n    }\n    sort {\n      id\n      order\n    }\n  }\n}\n',
                'variables': {
                    'filtersVersion': 4,
                    'query': str(scrapSKU),
                    'sort': {
                        'id': 'featured',
                        'order': 'DESC',
                    },
                    'staticRanking': {
                        'enabled': False,
                    },
                    'page': {
                        'index': 1,
                        'limit': 10,
                    },
                },
                'operationName': 'GetSearchResults',
            }
            url = 'https://stockx.com/api/p/e'
            response = req(session.post, url=url, json=json_data, proxies=PROX.get_current())
            need = response.json()
            search_total = need['data']['browse']['results']['pageInfo']['total']

            if search_total > 0 :
                primTitle = need['data']['browse']['results']['edges'][0]['node']['primaryTitle']
                secondaryTitle = need['data']['browse']['results']['edges'][0]['node']['secondaryTitle']
                urlKey = "https://stockx.com/" + need['data']['browse']['results']['edges'][0]['node']['urlKey']

                url_card = req(session.get, url=urlKey, proxies=PROX.get_current())

                soup = BeautifulSoup(url_card.text, "lxml")
                try: 
                    data = soup.find("p", class_="chakra-text css-wgsjnl").text
                except:
                    data = ''
                real_name = primTitle + " " + secondaryTitle + " " + data
                print(real_name)
                print(data)
                print(urlKey)

                excel_name = sheet[row][1]
                excel_name.value = real_name

                excel_sku = sheet[row][2]
                excel_sku.value = data

                excel_urlKey = sheet[row][3]
                excel_urlKey.value = urlKey

                book.save(filename)
                print("Файл обновил")
                PROX.next()    
            else:
                print ("товар не найден")
        else:
            print('Уже есть')
        await msg.edit_text(f'Выполнено {int((row-1)/sheet.max_row*100)}%. ({row-1}/{sheet.max_row-1})')
    print("Работа завершена")


def main():
    filename = 'sku.xlsx'
    find_goods(filename)


if __name__ == '__main__':
    main()